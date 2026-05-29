import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Power Query parameters (from "Receive Payments (Y&S Ledger).xlsm")
# ---------------------------------------------------------------------------

# Table1 query keeps rows whose Type is exactly one of these six strings.
RECEIPT_TYPES = {"Payment", "Refund", "payment", "refund", "Recoup", "recoup"}
DEPOSIT_TO = "Batch Payments Held by Affiliates"
DEPOSIT_TO_TV = "Batch Payments Held by TV"
EVENT_COLS = ["Performer", "Venue", "EventDate", "Section", "Row", "Seat", "Qty", "Reason"]

# Raw ledger schema (Table1)
RAW_COLS = ["Company", "Date", "Name", "Type", "Order #", "Amount"] + EVENT_COLS

# Received-payments output schema (Table1_2). Company holds the affiliate name.
RECEIPT_COLS = ["Company", "Customer", "Payment Date", "Deposit To", "Amount Received"]

# ---------------------------------------------------------------------------
# Master Mapping List  (ledger / TicketVault company -> QBO company)
# Hardcoded from Master_Mapping_List.xlsx, keeping ONLY rows whose
# "QBO Company" column is populated (rows marked N/A are excluded).
#   (ledger_company, qbo_company, applied_payments_category)
# ---------------------------------------------------------------------------
QBO_MAP_ROWS = [
    ("Damon and Crew",     "Damona & Crew",       "Affiliates"),
    ("The Ticket Guy",     "The Ticket Guy LLC",  "Other"),
    ("YS Tickets",         "Y&S Tickets",         "Y&S - RecPmt"),
    ("YS-SeatGeek2",       "Y&S Tickets",         "Y&S - RecPmt"),
    ("YS-Seatgeek",        "Y&S Tickets",         "Y&S - RecPmt"),
    ("YS Tickets Spec",    "Y&S Tickets",         "Y&S - RecPmt"),
    ("YourTickets",        "YourTickets",         "Affiliates"),
    ("YSA",                "YS Asher Tickets",    "Affiliates"),
    ("YSA 2",              "YS Asher Tickets",    "Affiliates"),
    ("YSA 3",              "YS Asher Tickets",    "Affiliates"),
    ("Jacks YS",           "YS Chase Tickets",    "Affiliates"),
    ("YS Katz",            "YS Katz Tickets",     "Affiliates"),
    ("Yoni Levine",        "YS Levine Tickets",   "Affiliates"),
    ("Levovitz",           "YS Levovitz Tickets", "Affiliates"),
    ("Needle Tickets LLC", "YS Needle Tickets",   "Affiliates"),
    ("YS TL",              "YS TL Tickets",       "Affiliates"),
    ("GK LLC",             "YSKG Tickets",        "Affiliates"),
    ("YSM Tickets",        "YSM Tickets",         "Affiliates"),
    ("Pollak Tickets",     "YSP Tickets",         "Affiliates"),
    ("YSS Tickets",        "YSS Tickets",         "Affiliates"),
    ("YSW",                "YSW Tickets",         "Affiliates"),
]

# Normalized lookup (trim + casefold) so minor casing/space differences match.
_QBO_LOOKUP = {tv.strip().lower(): qbo for tv, qbo, _ in QBO_MAP_ROWS}
QBO_CATEGORY = {qbo: cat for _, qbo, cat in QBO_MAP_ROWS}
# Distinct QBO companies, in first-appearance order.
QBO_COMPANIES = list(dict.fromkeys(qbo for _, qbo, _ in QBO_MAP_ROWS))

# QBO entities that should NOT get their own tab/file or appear in output.
EXCLUDE_QBO = {"Y&S Tickets"}


def _display_of_qbo(qbo):
    """Drop a trailing ' Tickets' for the short label used on tabs/files/data."""
    if isinstance(qbo, str) and qbo.endswith(" Tickets"):
        return qbo[:-len(" Tickets")]
    return qbo


# Affiliate output companies (short names), in workbook order.
OUTPUT_COMPANIES = [_display_of_qbo(q) for q in QBO_COMPANIES if q not in EXCLUDE_QBO]


def map_qbo(company_series):
    """Map a ledger Company series to QBO company names (unmapped -> NaN)."""
    return company_series.astype("string").str.strip().str.lower().map(_QBO_LOOKUP)


def map_company(company_series):
    """Map a ledger Company series to the short affiliate label (unmapped -> NaN)."""
    return map_qbo(company_series).map(_display_of_qbo)


def _strip_cad(s):
    """Remove a 'CAD' currency tag whether or not it's parenthesized.

    Matches whole-word CAD only (case-insensitive), with optional surrounding
    parentheses, so names like 'Arcadia' or 'Cadence' are never clipped.
    """
    s = re.sub(r"\(?\bCAD\b\)?", "", str(s), flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def _clean_customer(v):
    """Remove any 'CAD' tag, collapse spaces, then append ' (C)'."""
    if pd.isna(v):
        return v
    return f"{_strip_cad(v)} (C)"


DATE_COLS = {"Payment Date", "Date", "EventDate"}
AMOUNT_COLS = {"Amount Received", "Amount"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fix_date(col):
    """Coerce any date column to a real datetime64 series.

    Handles Excel serial numbers, date strings in mixed formats (ISO and
    M/D/YYYY), existing Timestamps, blanks, and unparseable/out-of-range
    values (which become NaT). Always returns a datetime64 series so .dt
    accessors are safe downstream.
    """
    s = pd.Series(col)
    idx = s.index
    s = s.reset_index(drop=True)

    # Already datetime — just normalize the dtype.
    if pd.api.types.is_datetime64_any_dtype(s):
        out = pd.to_datetime(s, errors="coerce")
        out.index = idx
        return out

    # Numeric Excel serials (also catches numbers stored as strings).
    num = pd.to_numeric(s, errors="coerce")
    serial_mask = num.notna() & (num > 59) & (num < 80000)

    non_serial = s.where(~serial_mask)
    if non_serial.notna().any():
        try:
            out = pd.to_datetime(non_serial, errors="coerce", format="mixed")
        except Exception:
            out = pd.to_datetime(non_serial, errors="coerce")
    else:
        out = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    if serial_mask.any():
        serials = pd.to_datetime(num[serial_mask], unit="D",
                                 origin="1899-12-30", errors="coerce")
        out.loc[serial_mask] = serials

    out.index = idx
    return out


def ordinal(n):
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def format_date_range(dates):
    if not dates:
        return "No Dates"
    sd = sorted(dates)
    fmt = lambda d: f"{d.strftime('%B')} {ordinal(d.day)}"
    if len(sd) == 1:
        return f"{fmt(sd[0])} {sd[0].strftime('%Y')}"
    return f"{fmt(sd[0])} thru {fmt(sd[-1])} {sd[-1].strftime('%Y')}"


def _to_num(s):
    return pd.to_numeric(s, errors="coerce")


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

def write_sheet(wb, name, dataframe):
    safe = name[:31]
    for ch in '[]:*?/\\':
        safe = safe.replace(ch, " ")
    ws = wb.create_sheet(safe)
    cols = list(dataframe.columns)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, border

    fill_odd = PatternFill("solid", start_color="FFFFFF")
    fill_even = PatternFill("solid", start_color="EEF2FF")
    body_font = Font(name="Arial", size=10)

    for ri, row in enumerate(dataframe.itertuples(index=False), 2):
        row_fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, val in enumerate(row, 1):
            col_name = cols[ci - 1]
            if col_name in DATE_COLS and val is not None and not pd.isna(val):
                try:
                    val = pd.Timestamp(val).strftime("%m/%d/%Y")
                except Exception:
                    pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font
            c.alignment = Alignment(vertical="center")
            c.border = border
            c.fill = row_fill
            if col_name in AMOUNT_COLS:
                c.number_format = "#,##0.00"

    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        for row in dataframe.itertuples(index=False):
            v = row[ci - 1]
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"
    if dataframe.shape[0] >= 0:
        ws.auto_filter.ref = ws.dimensions
    return ws


def add_company_summary(ws, dataframe, title="Summary by Company"):
    """Write a pivot-style 'Sum of Amount by Company' block to the right of the data."""
    if "Company" not in dataframe.columns or "Amount" not in dataframe.columns:
        return
    summary = (dataframe.groupby("Company", dropna=False)["Amount"].sum()
               .round(2).abs().sort_index())
    grand = round(float(summary.sum()), 2)

    sc = len(dataframe.columns) + 2          # leave one blank gap column
    L = get_column_letter(sc)
    R = get_column_letter(sc + 1)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color="4472C4")
    halign = Alignment(horizontal="center", vertical="center")

    # Title
    tcell = ws.cell(row=1, column=sc, value=title)
    tcell.font = Font(name="Arial", bold=True, size=11)
    # Header row
    for col, label in ((sc, "Company"), (sc + 1, "Sum of Amount")):
        c = ws.cell(row=2, column=col, value=label)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, border

    r = 3
    for company, amount in summary.items():
        cc = ws.cell(row=r, column=sc, value=("" if pd.isna(company) else company))
        ac = ws.cell(row=r, column=sc + 1, value=float(amount) if pd.notna(amount) else 0.0)
        cc.font = Font(name="Arial", size=10)
        ac.font = Font(name="Arial", size=10)
        ac.number_format = "#,##0.00"
        cc.border = ac.border = border
        r += 1

    # Grand total
    top = Side(style="thin", color="000000")
    gc = ws.cell(row=r, column=sc, value="Grand Total")
    ga = ws.cell(row=r, column=sc + 1, value=grand)
    gc.font = ga.font = Font(name="Arial", bold=True, size=10)
    ga.number_format = "#,##0.00"
    gc.border = ga.border = Border(top=top, bottom=Side(style="double", color="000000"))

    ws.column_dimensions[L].width = max(16, summary.index.map(lambda x: len(str(x))).max() + 2 if len(summary) else 16)
    ws.column_dimensions[R].width = 16


# ---------------------------------------------------------------------------
# Query reimplementations (mimicking the workbook's M code)
# ---------------------------------------------------------------------------

def build_received_payments(raw, deposit_to=DEPOSIT_TO, map_companies=True):
    """Consolidated Amount Received per company/customer/date.

    map_companies=True  -> map the source Company to its QBO affiliate name and
                           drop unmapped/excluded entities (Y&S ledger).
    map_companies=False -> keep Company as-is (TicketVault: it's the broker).
    """
    df = raw.copy()
    df["Company"] = df["Company"].astype("string")
    df["Date"] = fix_date(df["Date"])
    df["Name"] = df["Name"].astype("string")
    df["Type"] = df["Type"].astype("string")
    df["Amount"] = _to_num(df["Amount"])

    df = df[~df["Name"].isin(["Starting Balance", "Y&S"])]
    df = df[df["Type"].isin(RECEIPT_TYPES)]

    grouped = (
        df.groupby(["Company", "Name", "Date"], sort=False, dropna=False)["Amount"]
        .sum(min_count=1)
        .reset_index()
        .rename(columns={"Name": "Customer", "Amount": "Amount Received"})
    )

    # Map to affiliate label; drop unmapped (N/A) and excluded entities.
    if map_companies:
        qbo = map_qbo(grouped["Company"])
        keep = qbo.notna() & ~qbo.isin(EXCLUDE_QBO)
        grouped = grouped[keep].copy()
        grouped["Company"] = qbo[keep].map(_display_of_qbo).values

    grouped["Date"] = grouped["Date"].dt.normalize()
    grouped = grouped.rename(columns={"Date": "Payment Date"})
    grouped["Deposit To"] = deposit_to

    # Drop rows with no amount, strip "CAD" tag from customers.
    grouped = grouped[grouped["Amount Received"].notna()]
    grouped["Customer"] = grouped["Customer"].apply(
        lambda v: _strip_cad(v) if pd.notna(v) else v
    )

    # Consolidate (collapses CAD/non-CAD and YSA/YSA2/YSA3-style merges).
    grouped = (
        grouped.groupby(["Company", "Customer", "Payment Date", "Deposit To"],
                        sort=False, dropna=False)["Amount Received"]
        .sum(min_count=1)
        .reset_index()
    )
    grouped["Customer"] = grouped["Customer"].apply(
        lambda v: f"{v} (C)" if pd.notna(v) else v)
    grouped["Amount Received"] = grouped["Amount Received"].round(2)
    return grouped[RECEIPT_COLS]


def _ledger_by_type(raw, type_value, map_companies=True):
    df = raw.copy()
    df["Company"] = df["Company"].astype("string")
    df["Date"] = fix_date(df["Date"])
    df["Name"] = df["Name"].astype("string")
    df["Type"] = df["Type"].astype("string")
    df["Amount"] = _to_num(df["Amount"])

    df = df[df["Type"] == type_value]
    df = df[df["Amount"].notna()]
    df["Company"] = map_company(df["Company"]) if map_companies else df["Company"]
    df["Name"] = df["Name"].apply(_clean_customer)
    df["Date"] = df["Date"].dt.normalize()
    out = df[["Company", "Date", "Name", "Type", "Amount"]].copy()
    out["Amount"] = out["Amount"].round(2)
    return out.reset_index(drop=True)


def build_transfers(raw, map_companies=True):
    return _ledger_by_type(raw, "Transfer", map_companies)


def build_fx(raw, map_companies=True):
    return _ledger_by_type(raw, "FX", map_companies)


def _grouped_fees(raw, type_value, map_companies=True):
    """Group a fee Type by Company/Date/Type, summing Amount."""
    df = raw.copy()
    df["Company"] = df["Company"].astype("string")
    df["Date"] = fix_date(df["Date"])
    df["Type"] = df["Type"].astype("string")
    df["Amount"] = _to_num(df["Amount"])

    df = df[df["Type"] == type_value]
    df = df[df["Amount"].notna()]
    df["Company"] = map_company(df["Company"]) if map_companies else df["Company"]
    df["Date"] = df["Date"].dt.normalize()
    grouped = (
        df.groupby(["Company", "Date", "Type"], sort=False, dropna=False)["Amount"]
        .sum(min_count=1)
        .reset_index()
    )
    grouped["Amount"] = grouped["Amount"].round(2)
    return grouped[["Company", "Date", "Type", "Amount"]]


def build_cancellation_fees(raw, map_companies=True):
    """Cancellation Fees grouped by Company/Date/Type."""
    return _grouped_fees(raw, "Cancellation Fees", map_companies)


def build_transferless_link_fees(raw, map_companies=True):
    """Transferless Link Fees (TV format) grouped by Company/Date/Type."""
    return _grouped_fees(raw, "Transferless Link Fees", map_companies)


def build_other_fees(raw, map_companies=True):
    """Catch-all: every Type that isn't a receipt, FX, or Cancellation Fee.

    Groups by Company/Date/Type so distinct fee types (Transferless Link Fees,
    Transfer, etc.) stay separated within the one tab.
    """
    df = raw.copy()
    df["Company"] = df["Company"].astype("string")
    df["Date"] = fix_date(df["Date"])
    df["Type"] = df["Type"].astype("string")
    df["Amount"] = _to_num(df["Amount"])

    excluded = set(RECEIPT_TYPES) | {"FX", "Cancellation Fees"}
    df = df[~df["Type"].isin(excluded)]
    df = df[df["Amount"].notna() & df["Type"].notna()]
    df["Company"] = map_company(df["Company"]) if map_companies else df["Company"]
    df["Date"] = df["Date"].dt.normalize()
    grouped = (
        df.groupby(["Company", "Date", "Type"], sort=False, dropna=False)["Amount"]
        .sum(min_count=1)
        .reset_index()
    )
    grouped["Amount"] = grouped["Amount"].round(2)
    return grouped[["Company", "Date", "Type", "Amount"]]


def company_slice(receipts, company):
    """Per-affiliate slice of the received-payments table."""
    return receipts[receipts["Company"] == company].reset_index(drop=True)


# ---------------------------------------------------------------------------
# File loading
# ---------------------------------------------------------------------------

# Column aliases: maps alternate source headers (e.g. the TicketVault "TV"
# export) to the canonical schema. Applied only when the canonical column
# isn't already present.
COLUMN_ALIASES = {
    "Amount (QBO)": "Amount",
    "Payout": "Amount",
    "Network": "Name",
    "CancellationReason": "Reason",
}


def _normalize_columns(df):
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df.columns = [str(c).strip() for c in df.columns]
    for alias, canonical in COLUMN_ALIASES.items():
        if alias in df.columns and canonical not in df.columns:
            df = df.rename(columns={alias: canonical})
    return df


def load_file(file_bytes, filename=""):
    """Load one ledger file (xlsx/xlsm/csv) and normalize to the raw schema.

    Handles both the original ledger layout and the TicketVault "TV" export
    layout (Network -> Name, Amount (QBO) -> Amount, CancellationReason ->
    Reason) via COLUMN_ALIASES.
    """
    fname = filename.lower()
    if fname.endswith(".csv"):
        df = _normalize_columns(pd.read_csv(io.BytesIO(file_bytes)))
    else:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        # Pick the sheet that best matches the ledger schema (after aliasing).
        chosen, best = xl.sheet_names[0], -1
        for s in xl.sheet_names:
            head = _normalize_columns(xl.parse(s, nrows=0))
            score = sum(c in head.columns for c in ["Company", "Date", "Name", "Type", "Amount"])
            if score > best:
                best, chosen = score, s
        df = _normalize_columns(xl.parse(chosen))

    for col in RAW_COLS:
        if col not in df.columns:
            df[col] = None
    return df


# ---------------------------------------------------------------------------
# Main entry points (interface expected by app.py)
# ---------------------------------------------------------------------------

def process_file(file_bytes, filename=""):
    return process_files([(file_bytes, filename)])


def process_files(file_list, ledger_type="yns", broker=None):
    """Dispatch on ledger type. 'yns' = original Y&S behavior; 'tv' = TicketVault."""
    if str(ledger_type).lower() in ("tv", "ticketvault", "ticketvault ledger"):
        return _process_tv(file_list, broker)
    return _process_yns(file_list)


def _process_yns(file_list):
    frames = [load_file(fb, fn) for fb, fn in file_list]
    df_raw = pd.concat(frames, ignore_index=True)

    # Exclude rows with no amount.
    df_raw["Amount"] = pd.to_numeric(df_raw["Amount"], errors="coerce")
    df_raw = df_raw[df_raw["Amount"].notna()].reset_index(drop=True)

    # Exclude unmapped companies and excluded entities from ALL output.
    qbo_raw = map_qbo(df_raw["Company"])
    df_raw = df_raw[qbo_raw.notna() & ~qbo_raw.isin(EXCLUDE_QBO)].reset_index(drop=True)

    receipts = build_received_payments(df_raw)
    fx = build_fx(df_raw)
    cancellations = build_cancellation_fees(df_raw)
    transferless = build_transferless_link_fees(df_raw)

    # Split out rows whose aggregated Amount Received is negative. These go
    # ONLY on the cross-company "Negative Payments" tab, never on the company
    # tabs or individual files.
    positive = receipts[receipts["Amount Received"] >= 0].reset_index(drop=True)
    negative = receipts[receipts["Amount Received"] < 0].reset_index(drop=True)

    company_dfs = {name: company_slice(positive, name) for name in OUTPUT_COMPANIES}

    # Date range from all output dates (full receipts + fx + cancellation).
    dates = set()
    for col, frame in [("Payment Date", receipts), ("Date", fx), ("Date", cancellations)]:
        vals = pd.to_datetime(frame[col], errors="coerce").dropna()
        vals = vals[(vals.dt.year >= 2000) & (vals.dt.year <= 2100)]
        dates.update(d.date() for d in vals)
    date_range_str = format_date_range([pd.Timestamp(d) for d in dates])

    # ── Combined multi-sheet workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Input", df_raw[RAW_COLS])
    if len(negative):
        write_sheet(wb, "Negative Payments", negative)
    if len(fx):
        ws = write_sheet(wb, "FX", fx)
        add_company_summary(ws, fx)
    if len(cancellations):
        ws = write_sheet(wb, "Cancellation Fees", cancellations)
        add_company_summary(ws, cancellations)
    if len(transferless):
        ws = write_sheet(wb, "Transferless Link Fees", transferless)
        add_company_summary(ws, transferless)
    for name, cdf in company_dfs.items():
        write_sheet(wb, name, cdf)
    buf = io.BytesIO()
    wb.save(buf)
    combined_bytes = buf.getvalue()

    # ── One workbook per company (non-empty only, positives only) ──────────
    company_files = {}
    for name, cdf in company_dfs.items():
        if len(cdf) == 0:
            continue
        cwb = openpyxl.Workbook()
        cwb.remove(cwb.active)
        write_sheet(cwb, name, cdf)
        cbuf = io.BytesIO()
        cwb.save(cbuf)
        company_files[name] = cbuf.getvalue()

    # ── Per-company category totals for the summary table ──────────────────
    pos_by_co = positive.groupby("Company")["Amount Received"].sum()
    neg_by_co = negative.groupby("Company")["Amount Received"].sum()
    canc_by_co = cancellations.groupby("Company")["Amount"].sum().abs() if len(cancellations) else pd.Series(dtype=float)
    fx_by_co = fx.groupby("Company")["Amount"].sum().abs() if len(fx) else pd.Series(dtype=float)

    def g(series, key):
        try:
            v = float(series.get(key, 0.0))
        except Exception:
            v = 0.0
        return round(v, 2)

    stats = {}
    for name, cdf in company_dfs.items():
        stats[name] = {
            "rows": len(cdf),
            "receive": g(pos_by_co, name),
            "negative": g(neg_by_co, name),
            "cancellation": g(canc_by_co, name),
            "fx": g(fx_by_co, name),
        }
    stats["Combined"] = {
        "rows": int(sum(len(c) for c in company_dfs.values())),
        "receive": round(float(positive["Amount Received"].sum()), 2) if len(positive) else 0.0,
        "negative": round(float(negative["Amount Received"].sum()), 2) if len(negative) else 0.0,
        "cancellation": round(float(canc_by_co.sum()), 2) if len(canc_by_co) else 0.0,
        "fx": round(float(fx_by_co.sum()), 2) if len(fx_by_co) else 0.0,
    }

    return {
        "mode": "yns",
        "date_range": date_range_str,
        "combined": combined_bytes,
        "companies": company_files,
        "all_companies": list(company_dfs.keys()),
        "stats": stats,
    }


def _process_tv(file_list, broker):
    broker = (broker or "").strip()
    if not broker:
        raise ValueError("A broker is required for TicketVault ledgers.")

    frames = [load_file(fb, fn) for fb, fn in file_list]
    df_raw = pd.concat(frames, ignore_index=True)

    df_raw["Amount"] = pd.to_numeric(df_raw["Amount"], errors="coerce")
    df_raw = df_raw[df_raw["Amount"].notna()].reset_index(drop=True)

    # Stamp every row with the broker (TV files have no usable Company column).
    df_raw["Company"] = broker

    receipts = build_received_payments(df_raw, deposit_to=DEPOSIT_TO_TV, map_companies=False)
    fx = build_fx(df_raw, map_companies=False)
    cancellations = build_cancellation_fees(df_raw, map_companies=False)
    other = build_other_fees(df_raw, map_companies=False)

    positive = receipts[receipts["Amount Received"] >= 0].reset_index(drop=True)
    negative = receipts[receipts["Amount Received"] < 0].reset_index(drop=True)

    dates = set()
    for col, frame in [("Payment Date", receipts), ("Date", fx), ("Date", cancellations), ("Date", other)]:
        if len(frame):
            vals = pd.to_datetime(frame[col], errors="coerce").dropna()
            vals = vals[(vals.dt.year >= 2000) & (vals.dt.year <= 2100)]
            dates.update(d.date() for d in vals)
    date_range_str = format_date_range([pd.Timestamp(d) for d in dates])

    # ── File 1: Receive Payments (positives only) ──────────────────────────
    rp_wb = openpyxl.Workbook()
    rp_wb.remove(rp_wb.active)
    write_sheet(rp_wb, "Receive Payments", positive)
    rp_buf = io.BytesIO()
    rp_wb.save(rp_buf)
    receive_bytes = rp_buf.getvalue()

    # ── File 2: Everything ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Input", df_raw[RAW_COLS])
    write_sheet(wb, "Receive Payments", positive)
    if len(negative):
        write_sheet(wb, "Negative Payments", negative)
    if len(fx):
        ws = write_sheet(wb, "FX", fx)
        add_company_summary(ws, fx)
    if len(cancellations):
        ws = write_sheet(wb, "Cancellation Fees", cancellations)
        add_company_summary(ws, cancellations)
    if len(other):
        ws = write_sheet(wb, "Other Fees", other)
        add_company_summary(ws, other)
    everything_buf = io.BytesIO()
    wb.save(everything_buf)
    everything_bytes = everything_buf.getvalue()

    rcv = round(float(positive["Amount Received"].sum()), 2) if len(positive) else 0.0
    neg = round(float(negative["Amount Received"].sum()), 2) if len(negative) else 0.0
    canc = round(float(cancellations["Amount"].sum()), 2) if len(cancellations) else 0.0
    fxs = round(float(fx["Amount"].sum()), 2) if len(fx) else 0.0
    oth = round(float(other["Amount"].sum()), 2) if len(other) else 0.0

    return {
        "mode": "tv",
        "broker": broker,
        "date_range": date_range_str,
        "combined": everything_bytes,                 # the "Everything" file
        "companies": {"Receive Payments": receive_bytes},
        "all_companies": ["Receive Payments"],
        "stats": {"receive": rcv, "negative": neg, "cancellation": canc, "fx": fxs, "other": oth},
    }
